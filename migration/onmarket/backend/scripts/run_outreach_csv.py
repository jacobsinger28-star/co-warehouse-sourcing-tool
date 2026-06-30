"""Run the outreach engine over a CSV/Excel list of real leads.

This is the connective tissue between "engine works on a hardcoded sample" and
"engine produces the gating coverage number on *your* data." The workflow it
enables:

    1. Put leads in a spreadsheet (one row per contactable person).
    2. Run this script.
    3. Read the coverage report + the enriched output CSV.

Channel-agnostic on purpose — owner rows, attorney rows, or a mix all flow
through the same waterfall. Decide the channel per-row via the `channel`
column ("owner" | "attorney").

USAGE
-----
    cd backend

    # 1. Get a blank input template with the right headers:
    ./.venv/bin/python -m scripts.run_outreach_csv --template my_leads.csv

    # 2. Fill it in (Excel/Sheets), then run it:
    ./.venv/bin/python -m scripts.run_outreach_csv --in my_leads.csv

    # Try it right now on the bundled sample:
    ./.venv/bin/python -m scripts.run_outreach_csv --in scripts/sample_leads.csv

LIVE vs STUB
------------
Runs against deterministic stubs until provider keys are in the env. To get the
*real* gating coverage number, set APOLLO_API_KEY (and later SKIPTRACE_API_KEY)
and re-run. When Apollo is live, each row costs Apollo credits — use --limit to
test on a slice first. ANTHROPIC_API_KEY switches drafting from template to
claude-opus-4-8.

INPUT COLUMNS
-------------
Headers are matched case-insensitively; unknown columns are ignored and missing
ones default to empty. Real-world inputs are ragged and that's fine — give the
waterfall whatever you have. The fields that actually move coverage:
    name              first + last; Apollo's primary key
    company / domain   company web domain is Apollo's *best* key (esp. attorneys)
    mailing_address    the skip-trace key for private owners behind LLCs
    seed_email         if you already have it, the waterfall short-circuits
See OutreachLead in outreach/models.py for the full field list.
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import sys
from dataclasses import asdict, fields
from pathlib import Path

from outreach import OutreachLead, run_outreach, coverage_report
from outreach.models import CHANNEL_OWNER


# Every OutreachLead field is a valid input column. A few friendly aliases map
# common spreadsheet headers onto the canonical field names.
_LEAD_FIELDS = [f.name for f in fields(OutreachLead)]
_ALIASES = {
    "full_name": "name",
    "owner_name": "name",
    "attorney_name": "name",
    "firm": "company",
    "website": "domain",
    "email": "seed_email",
    "linkedin": "linkedin_url",
    "property": "property_address",
    "notes": "property_context",
    "llc": "llc_name",
}


def _canonical(header: str) -> str | None:
    """Map a spreadsheet header to an OutreachLead field name, or None."""
    key = (header or "").strip().lower().replace(" ", "_").replace("-", "_")
    if key in _LEAD_FIELDS:
        return key
    return _ALIASES.get(key)


def _row_to_lead(row: dict) -> OutreachLead:
    kwargs: dict = {}
    for header, value in row.items():
        field_name = _canonical(header)
        if field_name and value is not None and str(value).strip():
            kwargs[field_name] = str(value).strip()
    kwargs.setdefault("channel", CHANNEL_OWNER)
    return OutreachLead(**kwargs)


def _read_rows(path: Path) -> list[dict]:
    """Read CSV or XLSX into a list of {header: value} dicts."""
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit(
                "Reading .xlsx needs openpyxl (`pip install openpyxl`), or export "
                "the sheet to CSV and pass that instead."
            )
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = [str(h) if h is not None else "" for h in next(rows_iter)]
        except StopIteration:
            return []
        return [dict(zip(headers, r)) for r in rows_iter]
    with path.open(newline="", encoding="utf-8-sig") as fh:
        return list(csv.DictReader(fh))


# Output columns: lead fields, then enrichment, then draft. One row per lead.
_OUT_HEADERS = (
    _LEAD_FIELDS
    + ["enr_email", "enr_phone", "enr_linkedin", "enr_source", "enr_confidence", "enr_live"]
    + ["draft_subject", "draft_body", "draft_hook", "draft_model"]
)


def _result_to_row(result) -> dict:
    lead = asdict(result.lead)
    e = result.enrichment
    d = result.draft
    lead.update(
        enr_email=e.email or "",
        enr_phone=e.phone or "",
        enr_linkedin=e.linkedin_url or "",
        enr_source=e.source or "",
        enr_confidence=f"{e.confidence:.2f}",
        enr_live=e.live,
        draft_subject=d.subject if d else "",
        draft_body=d.body if d else "",
        draft_hook=(d.hook or "") if d else "",
        draft_model=d.model_used if d else "",
    )
    return lead


def _write_template(path: Path) -> None:
    headers = _LEAD_FIELDS
    example = {
        "name": "Jane Owner",
        "company": "Owner Industrial LLC",
        "domain": "ownerindustrial.com",
        "mailing_address": "100 Main St, Springfield, IL 62701",
        "property_address": "200 Logistics Way, Springfield, IL",
        "property_context": "110,000 SF distribution, last sold 2015",
        "channel": "owner",
        "llc_name": "200 Logistics Holdings LLC",
        "source_record": "Sangamon County assessor parcel 00-000",
    }
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=headers)
        writer.writeheader()
        writer.writerow({h: example.get(h, "") for h in headers})
    print(f"Wrote input template -> {path}")
    print("Fill one row per contactable person; delete the example row when done.")


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    parser.add_argument("--in", dest="in_path", help="Input CSV or XLSX of leads")
    parser.add_argument("--out", dest="out_path", help="Output enriched CSV "
                        "(default: <input>.enriched.csv)")
    parser.add_argument("--report", dest="report_path",
                        help="Optional path to write the coverage report as JSON")
    parser.add_argument("--template", dest="template_path",
                        help="Write a blank input template CSV here and exit")
    parser.add_argument("--no-draft", action="store_true",
                        help="Skip email drafting (faster; no Claude calls)")
    parser.add_argument("--limit", type=int, default=None,
                        help="Only process the first N rows (test Apollo credits cheaply)")
    args = parser.parse_args(argv)

    if args.template_path:
        _write_template(Path(args.template_path))
        return

    if not args.in_path:
        parser.error("one of --in or --template is required")

    in_path = Path(args.in_path)
    if not in_path.exists():
        sys.exit(f"Input file not found: {in_path}")

    rows = _read_rows(in_path)
    if args.limit is not None:
        rows = rows[: args.limit]
    leads = [_row_to_lead(r) for r in rows]
    if not leads:
        sys.exit(f"No lead rows found in {in_path}")

    apollo_live = bool(os.getenv("APOLLO_API_KEY"))
    if apollo_live:
        print(f"[LIVE] Apollo key detected — processing {len(leads)} rows will "
              f"consume ~{len(leads)}+ Apollo credits.")
    else:
        print(f"[STUB] No APOLLO_API_KEY — running {len(leads)} rows against stubs "
              f"(coverage numbers are not real yet).")

    results = run_outreach(leads, draft=not args.no_draft)

    out_path = Path(args.out_path) if args.out_path else in_path.with_suffix(".enriched.csv")
    with out_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=_OUT_HEADERS)
        writer.writeheader()
        for r in results:
            writer.writerow(_result_to_row(r))

    report = coverage_report(results)
    if args.report_path:
        Path(args.report_path).write_text(json.dumps(report, indent=2))

    print("\n=== COVERAGE REPORT ===")
    print(json.dumps(report, indent=2))
    if not report["providers_live"]:
        print("\n(! providers running as STUBS — set APOLLO_API_KEY / "
              "SKIPTRACE_API_KEY for the real gating coverage number)")
    print(f"\nEnriched rows -> {out_path}")


if __name__ == "__main__":
    main()
