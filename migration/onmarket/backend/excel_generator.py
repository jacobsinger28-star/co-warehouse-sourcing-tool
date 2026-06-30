"""
Generate Scored_Opportunities_[Date]_[Analyst].xlsx with conditional formatting.
"""
from __future__ import annotations
import io
from datetime import date

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILL_GREEN  = PatternFill(fill_type="solid", fgColor="C6EFCE")
FILL_YELLOW = PatternFill(fill_type="solid", fgColor="FFEB9C")
FILL_RED    = PatternFill(fill_type="solid", fgColor="FFC7CE")

FONT_GREEN  = Font(color="276221", bold=True)
FONT_YELLOW = Font(color="9C5700", bold=True)
FONT_RED    = Font(color="9C0006", bold=True)

HEADER_FILL  = PatternFill(fill_type="solid", fgColor="1F3864")
HEADER_FONT  = Font(color="FFFFFF", bold=True)
ANALYST_FILL = PatternFill(fill_type="solid", fgColor="EBF0FA")
ANALYST_FONT = Font(color="1F3864", bold=True, size=11)

THIN        = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ENRICHMENT_COLS = [
    "Score_Category",
    "Implied_Purchase_Price",
    "Power_Density",
    "Truck_Court_Depth",
    "Pricing_Delta",
    "Scoring_Reason",
]

COL_WIDTHS = {
    "address":                     38,
    "Score_Category":               16,
    "Implied_Purchase_Price":       22,
    "Power_Density":                20,
    "Truck_Court_Depth":            22,
    "Pricing_Delta":                16,
    "Scoring_Reason":               60,
    "market_gross_rent_small_bay":  24,
    "asking_price_psf":             18,
    "truck_court_depth":            20,
    "notes":                        35,
}


def generate_excel(scored_df: pd.DataFrame, analyst_name: str = "") -> tuple[bytes, str]:
    today     = date.today().strftime("%Y-%m-%d")
    safe_name = analyst_name.strip().replace(" ", "_") if analyst_name else ""
    filename  = f"Scored_Opportunities_{today}{'_' + safe_name if safe_name else ''}.xlsx"

    original_cols = [c for c in scored_df.columns if c not in ENRICHMENT_COLS]
    ordered_cols  = original_cols + [c for c in ENRICHMENT_COLS if c in scored_df.columns]
    df = scored_df[ordered_cols]

    buffer = io.BytesIO()
    df.to_excel(buffer, index=False, sheet_name="Scored Opportunities", engine="openpyxl", startrow=1)
    buffer.seek(0)

    wb = load_workbook(buffer)
    ws = wb.active

    # --- Analyst banner row ---
    ws.insert_rows(1)
    banner = f"Uploaded by: {analyst_name}" if analyst_name else "Easybay Sourcing Tool"
    cell = ws.cell(row=1, column=1, value=banner)
    cell.font      = ANALYST_FONT
    cell.fill      = ANALYST_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
    ws.row_dimensions[1].height = 22

    # --- Header row (row 2) ---
    for cell in ws[2]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
    ws.row_dimensions[2].height = 30

    # --- Locate Score_Category column ---
    headers = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
    try:
        score_col = headers.index("Score_Category") + 1
    except ValueError:
        score_col = None

    # --- Format header labels for financial columns ---
    financial_labels = {
        "Implied_Purchase_Price": "Implied Price\n($/SF)",
        "Pricing_Delta":          "Pricing Delta\n($/SF)",
        "Truck_Court_Depth":      "Truck Court\n(Target: 100')",
        "Power_Density":          "Power Density",
        "Scoring_Reason":         "Scoring Reason",
    }
    for i, col_name in enumerate(ordered_cols, start=1):
        if col_name in financial_labels:
            ws.cell(row=2, column=i).value = financial_labels[col_name]

    # --- Data rows ---
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        category = row[score_col - 1].value if score_col else None
        if category == "Actionable":
            fill, font = FILL_GREEN,  FONT_GREEN
        elif category == "Tentative":
            fill, font = FILL_YELLOW, FONT_YELLOW
        elif category == "Pass":
            fill, font = FILL_RED,    FONT_RED
        else:
            fill, font = None, None

        for cell in row:
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            # Color the whole row lightly, Score_Category cell bold
            if fill:
                if cell.column == score_col:
                    cell.fill = fill
                    cell.font = font
                else:
                    # Light tint on all cells in the row
                    light_fill = PatternFill(
                        fill_type="solid",
                        fgColor=fill.fgColor.rgb if hasattr(fill.fgColor, 'rgb') else fill.fgColor.value
                    )
                    cell.fill = light_fill

    # Format $/SF columns as currency-ish
    for i, col_name in enumerate(ordered_cols, start=1):
        col_letter = get_column_letter(i)
        if col_name in ("Implied_Purchase_Price", "Pricing_Delta",
                        "market_gross_rent_small_bay", "asking_price_psf"):
            for row_num in range(3, ws.max_row + 1):
                c = ws.cell(row=row_num, column=i)
                if isinstance(c.value, (int, float)):
                    c.number_format = '"$"#,##0.00'

    # --- Column widths ---
    for i, col_name in enumerate(ordered_cols, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = COL_WIDTHS.get(col_name, 18)

    # --- Freeze & filter ---
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), filename
